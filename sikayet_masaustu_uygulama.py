"""
Sikayet Takip Sistemi - Masaustu Uygulamasi
----------------------------------------------
Bu, cift tiklayip acilan basit bir pencere uygulamasidir (tkinter ile,
Python'un icinde hazir gelir, ayrica kurulum gerektirmez).

ONEMLI ON KOSUL:
    Bu uygulamanin calistigi bilgisayarda Tesseract OCR programinin
    kurulu olmasi gerekir (bkz. daha once konustugumuz kurulum adimlari).
    Tesseract kurulu degilse, uygulama acilir ama "Isle" butonuna
    bastiginda hata verir.

KURULUM:
    pip install pytesseract openpyxl pillow

CALISTIRMA:
    python sikayet_masaustu_uygulama.py
"""

import re
import os
import threading
from pathlib import Path
from tkinter import (
    Tk, Frame, Label, Button, Listbox, Scrollbar, END, filedialog, messagebox, StringVar
)

import pytesseract
from PIL import Image
from openpyxl import Workbook, load_workbook

# Windows'ta Tesseract'in kurulu oldugu yol farkliysa burayi guncelle.
# Mac/Linux'ta bu satiri yorum satiri yapabilirsin (# ile basiat).
if os.name == "nt":
    varsayilan_yol = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.path.exists(varsayilan_yol):
        pytesseract.pytesseract.tesseract_cmd = varsayilan_yol

# ---------------------- AYARLAR ----------------------
LANG = "tur+eng"
VALUE_COL_X = 140
RIGHT_COL_LIMIT = 500   # sagdaki ikinci sutuna (orn. "Sikayet") tasmayi engeller
ROW_TOLERANCE = 8
EXCEL_PATH = "sikayet_kayitlari.xlsx"
SHEET_HEADERS = ["Gorev No", "Ad Soyad", "Telefon", "Görev Tipi", "Şikayet Metni", "Kaynak Dosya"]
LABEL_KEYWORDS = {"ad_soyad": ["talep eden", "sms"], "telefon": ["telefon"]}
# -------------------------------------------------------


def gorev_no_bul(text):
    m = re.search(r"No\s*-?\s*\[?\s*(\d{4,})\s*\]?", text)
    return m.group(1) if m else ""


def satirlari_grupla(image_path):
    img = Image.open(image_path)
    data = pytesseract.image_to_data(img, lang=LANG, output_type=pytesseract.Output.DICT)
    kelimeler = []
    for i in range(len(data["text"])):
        t = data["text"][i].strip()
        if t:
            kelimeler.append((data["top"][i], data["left"][i], t))
    kelimeler.sort()
    satirlar = []
    for top, left, t in kelimeler:
        eklendi = False
        for satir in satirlar:
            if abs(satir["top"] - top) <= ROW_TOLERANCE:
                satir["items"].append((left, t))
                satir["top"] = (satir["top"] + top) / 2
                eklendi = True
                break
        if not eklendi:
            satirlar.append({"top": top, "items": [(left, t)]})
    satirlar.sort(key=lambda s: s["top"])
    return satirlar


def satirdan_etiket_deger_ayir(satir):
    items = sorted(satir["items"])
    etiket = " ".join(t for left, t in items if left < VALUE_COL_X)
    deger = " ".join(t for left, t in items if VALUE_COL_X <= left < RIGHT_COL_LIMIT)
    return etiket.lower(), deger


def deger_bul(satirlar, anahtar_kelimeler):
    for satir in satirlar:
        etiket, deger = satirdan_etiket_deger_ayir(satir)
        if any(k in etiket for k in anahtar_kelimeler) and deger:
            return deger.strip()
    return ""


def telefon_temizle(ham_deger):
    return re.sub(r"\D", "", ham_deger)


def ad_soyad_temizle(ham_deger):
    return re.sub(r"^[^\wÇĞİÖŞÜçğıöşü]+|[^\wÇĞİÖŞÜçğıöşü]+$", "", ham_deger).strip()


def gorev_tipi_ve_sikayet_metnini_ayikla(satirlar):
    """Gorev Tipi degerini ve Aciklama etiketinden once gelen paragrafi
    AYRI AYRI dondurur: (gorev_tipi, sikayet_metni)."""
    idx_tipi, idx_aciklama = None, None
    for i, satir in enumerate(satirlar):
        etiket, _ = satirdan_etiket_deger_ayir(satir)
        if idx_tipi is None and "tipi" in etiket:
            idx_tipi = i
        if idx_tipi is not None and "iklama" in etiket:
            idx_aciklama = i
            break

    gorev_tipi_degeri = deger_bul(satirlar, ["tipi"])
    sikayet_metni = ""
    if idx_tipi is not None and idx_aciklama is not None and idx_aciklama > idx_tipi:
        parcalar = []
        for satir in satirlar[idx_tipi + 1: idx_aciklama]:
            items = sorted(satir["items"])
            parcalar.append(" ".join(t for _, t in items))
        sikayet_metni = " ".join(p for p in parcalar if p).strip()

    return gorev_tipi_degeri, sikayet_metni


def gorseli_isle(image_path):
    img = Image.open(image_path)
    tum_metin = pytesseract.image_to_string(img, lang=LANG)
    satirlar = satirlari_grupla(image_path)
    ad_soyad = ad_soyad_temizle(deger_bul(satirlar, LABEL_KEYWORDS["ad_soyad"]))
    telefon_ham = deger_bul(satirlar, LABEL_KEYWORDS["telefon"])
    gorev_tipi, sikayet_metni = gorev_tipi_ve_sikayet_metnini_ayikla(satirlar)
    return {
        "Gorev No": gorev_no_bul(tum_metin),
        "Ad Soyad": ad_soyad,
        "Telefon": telefon_temizle(telefon_ham),
        "Görev Tipi": gorev_tipi,
        "Şikayet Metni": sikayet_metni,
        "Kaynak Dosya": Path(image_path).name,
    }


def excel_ac_veya_olustur(path):
    if Path(path).exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sikayetler"
        ws.append(SHEET_HEADERS)
    return wb, ws


class SikayetUygulamasi:
    def __init__(self, pencere):
        self.pencere = pencere
        pencere.title("Sikayet Takip Sistemi")
        pencere.geometry("560x420")

        self.secili_dosyalar = []

        ust_cerceve = Frame(pencere, pady=10)
        ust_cerceve.pack(fill="x")

        Label(ust_cerceve, text="Sikayet Takip Sistemi", font=("Segoe UI", 14, "bold")).pack()

        buton_cerceve = Frame(pencere, pady=10)
        buton_cerceve.pack(fill="x")

        Button(buton_cerceve, text="Gorsel Sec...", command=self.gorsel_sec, width=18).pack(side="left", padx=10)
        Button(buton_cerceve, text="Isle ve Excel'e Ekle", command=self.isle_baslat, width=20,
               bg="#4CAF50", fg="white").pack(side="left", padx=10)

        self.durum_metni = StringVar(value="Henuz gorsel secilmedi.")
        Label(pencere, textvariable=self.durum_metni, fg="gray").pack(pady=5)

        liste_cerceve = Frame(pencere)
        liste_cerceve.pack(fill="both", expand=True, padx=10, pady=10)

        Label(liste_cerceve, text="Islem Gecmisi:").pack(anchor="w")

        scrollbar = Scrollbar(liste_cerceve)
        scrollbar.pack(side="right", fill="y")

        self.liste_kutusu = Listbox(liste_cerceve, yscrollcommand=scrollbar.set)
        self.liste_kutusu.pack(fill="both", expand=True)
        scrollbar.config(command=self.liste_kutusu.yview)

        alt_cerceve = Frame(pencere, pady=10)
        alt_cerceve.pack(fill="x")
        Button(alt_cerceve, text="Excel Dosyasini Ac", command=self.excel_ac).pack(side="left", padx=10)

    def gorsel_sec(self):
        dosyalar = filedialog.askopenfilenames(
            title="Sikayet gorsellerini sec",
            filetypes=[("Gorsel dosyalari", "*.jpg *.jpeg *.png")]
        )
        if dosyalar:
            self.secili_dosyalar = list(dosyalar)
            self.durum_metni.set(f"{len(dosyalar)} gorsel secildi. Islemek icin butona bas.")

    def isle_baslat(self):
        if not self.secili_dosyalar:
            messagebox.showwarning("Uyari", "Once bir veya birden fazla gorsel sec.")
            return
        # Arayuzun donmamasi icin isleme ayri bir thread'de yapilir
        threading.Thread(target=self._isle, daemon=True).start()

    def _isle(self):
        self.durum_metni.set("Isleniyor, lutfen bekle...")
        wb, ws = excel_ac_veya_olustur(EXCEL_PATH)

        for dosya_yolu in self.secili_dosyalar:
            try:
                kayit = gorseli_isle(dosya_yolu)
            except Exception as e:
                self.liste_kutusu.insert(END, f"HATA ({Path(dosya_yolu).name}): {e}")
                continue

            ws.append([kayit["Gorev No"], kayit["Ad Soyad"], kayit["Telefon"], kayit["Görev Tipi"], kayit["Şikayet Metni"], kayit["Kaynak Dosya"]])

            satir = f"{kayit['Ad Soyad']} | {kayit['Telefon']} | {kayit['Kaynak Dosya']}"
            if not kayit["Ad Soyad"] or not kayit["Telefon"]:
                satir += "  [UYARI: eksik alan, kontrol et]"
            self.liste_kutusu.insert(END, satir)

        wb.save(EXCEL_PATH)
        self.secili_dosyalar = []
        self.durum_metni.set(f"Tamamlandi. Kayitlar '{EXCEL_PATH}' dosyasina yazildi.")

    def excel_ac(self):
        if not Path(EXCEL_PATH).exists():
            messagebox.showinfo("Bilgi", "Henuz olusturulmus bir Excel dosyasi yok.")
            return
        try:
            os.startfile(EXCEL_PATH)  # Windows
        except AttributeError:
            os.system(f'open "{EXCEL_PATH}"')  # Mac


if __name__ == "__main__":
    pencere = Tk()
    uygulama = SikayetUygulamasi(pencere)
    pencere.mainloop()
